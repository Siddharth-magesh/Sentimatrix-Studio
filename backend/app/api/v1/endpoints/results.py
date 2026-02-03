"""Results endpoints."""

from datetime import datetime
from typing import Annotated, Any

from fastapi import APIRouter, Depends, Query, status
from fastapi.responses import StreamingResponse

from app.core.deps import get_current_user
from app.models.result import (
    Result,
    ResultAggregation,
    ResultFilter,
    ResultList,
)
from app.models.user import User
from app.repositories.project import ProjectRepository, get_project_repository
from app.repositories.result import ResultRepository, get_result_repository

router = APIRouter()


@router.get(
    "/{project_id}/results",
    response_model=ResultList,
    summary="List project results",
)
async def list_results(
    project_id: str,
    current_user: Annotated[User, Depends(get_current_user)],
    project_repo: Annotated[ProjectRepository, Depends(get_project_repository)],
    result_repo: Annotated[ResultRepository, Depends(get_result_repository)],
    sentiment: str | None = Query(None, description="Filter by sentiment"),
    platform: str | None = Query(None, description="Filter by platform"),
    target_id: str | None = Query(None, description="Filter by target"),
    search: str | None = Query(None, description="Search in text/title"),
    date_from: datetime | None = Query(None, description="Results after this date"),
    date_to: datetime | None = Query(None, description="Results before this date"),
    page: int = Query(1, ge=1, description="Page number"),
    page_size: int = Query(20, ge=1, le=100, description="Items per page"),
) -> ResultList:
    """
    Get paginated list of analysis results for a project.

    Supports filtering by sentiment, platform, target, and date range.
    Use search to find specific text in results.
    """
    # Verify project ownership
    await project_repo.get_project(project_id, current_user.id)

    filters = ResultFilter(
        sentiment=sentiment,
        platform=platform,
        target_id=target_id,
        search=search,
        date_from=date_from,
        date_to=date_to,
    )

    return await result_repo.get_results_by_project(
        project_id=project_id,
        user_id=current_user.id,
        filters=filters,
        page=page,
        page_size=page_size,
    )


@router.get(
    "/{project_id}/results/{result_id}",
    response_model=Result,
    summary="Get result",
)
async def get_result(
    project_id: str,
    result_id: str,
    current_user: Annotated[User, Depends(get_current_user)],
    project_repo: Annotated[ProjectRepository, Depends(get_project_repository)],
    result_repo: Annotated[ResultRepository, Depends(get_result_repository)],
) -> Result:
    """
    Get a specific analysis result by ID.
    """
    # Verify project ownership
    await project_repo.get_project(project_id, current_user.id)

    return await result_repo.get_result(result_id, current_user.id)


@router.delete(
    "/{project_id}/results/{result_id}",
    status_code=status.HTTP_204_NO_CONTENT,
    summary="Delete result",
)
async def delete_result(
    project_id: str,
    result_id: str,
    current_user: Annotated[User, Depends(get_current_user)],
    project_repo: Annotated[ProjectRepository, Depends(get_project_repository)],
    result_repo: Annotated[ResultRepository, Depends(get_result_repository)],
) -> None:
    """
    Delete a specific result.
    """
    # Verify project ownership
    await project_repo.get_project(project_id, current_user.id)

    await result_repo.delete_result(result_id, current_user.id)


@router.get(
    "/{project_id}/results/analytics/summary",
    response_model=ResultAggregation,
    summary="Get results summary",
)
async def get_results_summary(
    project_id: str,
    current_user: Annotated[User, Depends(get_current_user)],
    project_repo: Annotated[ProjectRepository, Depends(get_project_repository)],
    result_repo: Annotated[ResultRepository, Depends(get_result_repository)],
    target_id: str | None = Query(None, description="Filter by target"),
) -> ResultAggregation:
    """
    Get aggregated statistics for project results.

    Returns total count, sentiment distribution, average scores, and date range.
    """
    # Verify project ownership
    await project_repo.get_project(project_id, current_user.id)

    return await result_repo.get_aggregation(
        project_id=project_id,
        user_id=current_user.id,
        target_id=target_id,
    )


@router.get(
    "/{project_id}/results/analytics/sentiment-timeline",
    response_model=list[dict],
    summary="Get sentiment over time",
)
async def get_sentiment_timeline(
    project_id: str,
    current_user: Annotated[User, Depends(get_current_user)],
    project_repo: Annotated[ProjectRepository, Depends(get_project_repository)],
    result_repo: Annotated[ResultRepository, Depends(get_result_repository)],
    target_id: str | None = Query(None, description="Filter by target"),
    interval: str = Query("day", description="Aggregation interval (hour, day, week, month)"),
) -> list[dict[str, Any]]:
    """
    Get sentiment scores over time.

    Returns average sentiment score for each time bucket, useful for trend analysis.
    """
    # Verify project ownership
    await project_repo.get_project(project_id, current_user.id)

    return await result_repo.get_sentiment_over_time(
        project_id=project_id,
        user_id=current_user.id,
        target_id=target_id,
        interval=interval,
    )


@router.get(
    "/{project_id}/results/analytics/emotions",
    response_model=list[dict],
    summary="Get top emotions",
)
async def get_top_emotions(
    project_id: str,
    current_user: Annotated[User, Depends(get_current_user)],
    project_repo: Annotated[ProjectRepository, Depends(get_project_repository)],
    result_repo: Annotated[ResultRepository, Depends(get_result_repository)],
    limit: int = Query(10, ge=1, le=50, description="Number of emotions to return"),
) -> list[dict[str, Any]]:
    """
    Get most frequently detected emotions across project results.
    """
    # Verify project ownership
    await project_repo.get_project(project_id, current_user.id)

    return await result_repo.get_top_emotions(
        project_id=project_id,
        user_id=current_user.id,
        limit=limit,
    )


@router.get(
    "/{project_id}/results/export",
    summary="Export results",
)
async def export_results(
    project_id: str,
    current_user: Annotated[User, Depends(get_current_user)],
    project_repo: Annotated[ProjectRepository, Depends(get_project_repository)],
    result_repo: Annotated[ResultRepository, Depends(get_result_repository)],
    format: str = Query("json", description="Export format (json, csv, xlsx)"),
) -> Any:
    """
    Export all results for a project.

    Supported formats:
    - **json**: JSON array of all results
    - **csv**: CSV file with flattened data
    - **xlsx**: Excel file with formatted data and multiple sheets
    """
    import csv
    import io
    import json

    # Verify project ownership
    project = await project_repo.get_project(project_id, current_user.id)

    results = await result_repo.export_results(
        project_id=project_id,
        user_id=current_user.id,
        format=format,
    )

    # Helper function to flatten results
    def flatten_results(results_data: list) -> list[dict]:
        flat_results = []
        for r in results_data:
            flat = {
                "id": r.get("id"),
                "platform": r.get("platform"),
                "text": r.get("content", {}).get("text"),
                "title": r.get("content", {}).get("title"),
                "author": r.get("content", {}).get("author"),
                "rating": r.get("content", {}).get("rating"),
                "date": r.get("content", {}).get("date"),
                "url": r.get("content", {}).get("url"),
                "sentiment_label": r.get("analysis", {}).get("sentiment", {}).get("label"),
                "sentiment_score": r.get("analysis", {}).get("sentiment", {}).get("score"),
                "sentiment_confidence": r.get("analysis", {}).get("sentiment", {}).get("confidence"),
                "primary_emotion": r.get("analysis", {}).get("emotions", {}).get("primary"),
                "emotion_scores": str(r.get("analysis", {}).get("emotions", {}).get("scores", {})),
                "created_at": r.get("created_at"),
            }
            flat_results.append(flat)
        return flat_results

    if format == "xlsx":
        # Excel export using openpyxl
        try:
            from openpyxl import Workbook
            from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
            from openpyxl.utils.dataframe import dataframe_to_rows
        except ImportError:
            from fastapi import HTTPException
            raise HTTPException(
                status_code=400,
                detail="Excel export requires openpyxl. Install with: pip install openpyxl",
            )

        wb = Workbook()

        # Sheet 1: All Results
        ws_results = wb.active
        ws_results.title = "Results"

        # Define styles
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        thin_border = Border(
            left=Side(style="thin"),
            right=Side(style="thin"),
            top=Side(style="thin"),
            bottom=Side(style="thin"),
        )

        # Headers
        headers = [
            "ID", "Platform", "Title", "Text", "Author", "Rating", "Date", "URL",
            "Sentiment", "Score", "Confidence", "Primary Emotion", "Created At"
        ]
        for col, header in enumerate(headers, 1):
            cell = ws_results.cell(row=1, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal="center")
            cell.border = thin_border

        # Data rows
        flat_results = flatten_results(results)
        for row_idx, result in enumerate(flat_results, 2):
            ws_results.cell(row=row_idx, column=1, value=result.get("id"))
            ws_results.cell(row=row_idx, column=2, value=result.get("platform"))
            ws_results.cell(row=row_idx, column=3, value=result.get("title"))
            ws_results.cell(row=row_idx, column=4, value=result.get("text", "")[:32000])  # Excel cell limit
            ws_results.cell(row=row_idx, column=5, value=result.get("author"))
            ws_results.cell(row=row_idx, column=6, value=result.get("rating"))
            ws_results.cell(row=row_idx, column=7, value=result.get("date"))
            ws_results.cell(row=row_idx, column=8, value=result.get("url"))
            ws_results.cell(row=row_idx, column=9, value=result.get("sentiment_label"))
            ws_results.cell(row=row_idx, column=10, value=result.get("sentiment_score"))
            ws_results.cell(row=row_idx, column=11, value=result.get("sentiment_confidence"))
            ws_results.cell(row=row_idx, column=12, value=result.get("primary_emotion"))
            ws_results.cell(row=row_idx, column=13, value=str(result.get("created_at", "")))

        # Auto-adjust column widths
        for col in ws_results.columns:
            max_length = 0
            column = col[0].column_letter
            for cell in col:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = min(len(str(cell.value)), 50)
                except Exception:
                    pass
            ws_results.column_dimensions[column].width = max_length + 2

        # Sheet 2: Summary Statistics
        ws_summary = wb.create_sheet("Summary")

        # Calculate summary stats
        total_results = len(results)
        sentiment_counts = {"positive": 0, "neutral": 0, "negative": 0}
        platform_counts = {}
        total_score = 0
        score_count = 0

        for r in results:
            sentiment = r.get("analysis", {}).get("sentiment", {}).get("label", "unknown")
            if sentiment in sentiment_counts:
                sentiment_counts[sentiment] += 1

            platform = r.get("platform", "unknown")
            platform_counts[platform] = platform_counts.get(platform, 0) + 1

            score = r.get("analysis", {}).get("sentiment", {}).get("score")
            if score is not None:
                total_score += score
                score_count += 1

        avg_score = total_score / score_count if score_count > 0 else 0

        # Write summary
        summary_data = [
            ["Metric", "Value"],
            ["Total Results", total_results],
            ["Average Sentiment Score", round(avg_score, 3)],
            ["Positive Results", sentiment_counts["positive"]],
            ["Neutral Results", sentiment_counts["neutral"]],
            ["Negative Results", sentiment_counts["negative"]],
            [""],
            ["Platform Breakdown", ""],
        ]

        for platform, count in platform_counts.items():
            summary_data.append([platform.title(), count])

        for row_idx, row_data in enumerate(summary_data, 1):
            for col_idx, value in enumerate(row_data, 1):
                cell = ws_summary.cell(row=row_idx, column=col_idx, value=value)
                if row_idx == 1:
                    cell.font = header_font
                    cell.fill = header_fill

        ws_summary.column_dimensions["A"].width = 25
        ws_summary.column_dimensions["B"].width = 15

        # Save to bytes
        output = io.BytesIO()
        wb.save(output)
        output.seek(0)

        return StreamingResponse(
            iter([output.getvalue()]),
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={
                "Content-Disposition": f"attachment; filename={project.name}_results.xlsx"
            },
        )

    if format == "csv":
        output = io.StringIO()
        if results:
            flat_results = flatten_results(results)
            writer = csv.DictWriter(output, fieldnames=flat_results[0].keys())
            writer.writeheader()
            writer.writerows(flat_results)

        return StreamingResponse(
            iter([output.getvalue()]),
            media_type="text/csv",
            headers={
                "Content-Disposition": f"attachment; filename={project.name}_results.csv"
            },
        )

    # Default JSON
    return StreamingResponse(
        iter([json.dumps(results, default=str, indent=2)]),
        media_type="application/json",
        headers={
            "Content-Disposition": f"attachment; filename={project.name}_results.json"
        },
    )


@router.delete(
    "/{project_id}/results",
    status_code=status.HTTP_204_NO_CONTENT,
    summary="Delete all results",
)
async def delete_all_results(
    project_id: str,
    current_user: Annotated[User, Depends(get_current_user)],
    project_repo: Annotated[ProjectRepository, Depends(get_project_repository)],
    result_repo: Annotated[ResultRepository, Depends(get_result_repository)],
    confirm: bool = Query(False, description="Confirm deletion"),
) -> None:
    """
    Delete all results for a project.

    Requires confirm=true to prevent accidental deletion.
    """
    if not confirm:
        from fastapi import HTTPException
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail="Set confirm=true to delete all results",
        )

    # Verify project ownership
    await project_repo.get_project(project_id, current_user.id)

    await result_repo.delete_results_by_project(project_id, current_user.id)
